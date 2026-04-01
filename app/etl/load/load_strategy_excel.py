import pandas as pd
from datetime import datetime
from etl.load.load_strategy import LoadStrategy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

class LoadStrategyExcel(LoadStrategy):
    """
    The responsability of this class is to load data in an excel
    """
    def load(df: pd.DataFrame, path_folder="etl/output"):    
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filepath = path_folder + "/job_offers_" + timestamp + ".xlsx"

        df.to_excel(filepath, index=False)
        
        # Add decoration
        LoadStrategyExcel.__decoration(filepath)

    def __decoration(filepath):
        wb = load_workbook(filepath)
        ws = wb.active

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", name="Arial")
        header_fill = PatternFill("solid", start_color="2F75B6")
        header_align = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Auto-fit column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 4

        wb.save(filepath)
